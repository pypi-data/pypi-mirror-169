from openpyxl import Workbook
from openpyxl.styles import PatternFill
from pathlib import Path
import re
# pylint: disable=no-member
import PySimpleGUIQt as sg

import data_minion.files_and_folders as ff
import data_minion.handle_settings as hs
import data_minion.split_students as ss

def get_grades(student: dict):
    grades = []
    for course in student['courses']:
        if course['course'].startswith('ORN010'):
            continue
        grade = course['grade']
        if grade.startswith('-'):
            continue
        grade = re.sub(r'%| |[A-Za-z]|-|\+', '', grade)
        try:
            grade = float(grade)
        except:
            continue
        grades.append(grade)
    return grades

def average_grade(grades: list):
    if grades != []:
        avg_grade = sum(grades) / len(grades)
        return round(avg_grade, 2)
    else:
        return None

def build_sheet(sheet, students):
    count = 2
    sheet.append(['Student Name', 'Course', 'Grade', 'Overdue', 'Last Visited'])
    header_fill = PatternFill(start_color="B8CCE4", fill_type = "solid")
    for student in students:
        avg_grade = average_grade(get_grades(student))
        summary_row = [student['name'], '', avg_grade]
        sheet.append(summary_row)
        count += 1
        first = count
        for course in student['courses']:
            row = [
                student['name'],
                course['course'],
                course['grade'],
                course['overdue'],
                course['visited'],
            ]
            sheet.append(row)
            count += 1
        end = count-1
        sheet.row_dimensions.group(first, end, hidden=True)
    for letter in 'ABCDEF':
        sheet[f'{letter}1'].fill = header_fill

def style_sheet(sheet):
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 8
    sheet.column_dimensions['F'].width = 25

def make_sheet(sheet, students):
    build_sheet(sheet, students)
    style_sheet(sheet)

##############################################################


def export_to_xlsx(fn, class_list: str):
    all_students = ff.get_all_student_info_from_file(fn, class_list)
    kudos, uhoh = ss.return_split_students(all_students)
    settings = hs.get_settings()
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'All Students'

    make_sheet(sheet, all_students)

    wb.create_sheet('Kudos')
    make_sheet(wb['Kudos'], kudos)

    wb.create_sheet('Uhoh')
    make_sheet(wb['Uhoh'], uhoh)
    exported_fn = sg.popup_get_file('', no_window=True, save_as=True, file_types=(('Excel Files', '*.xlsx'),), initial_folder=settings['xlsx_dir'])
    if not exported_fn:
        return 'cancelled'
    exported_fn = Path(exported_fn).absolute()
    xlsx_dir = exported_fn.as_posix()
    hs.edit_settings('xslx_dir', xlsx_dir)
    exported_fn = exported_fn.as_posix()
    try:
        wb.save(exported_fn)
        return exported_fn
    except PermissionError:
        return 'PermissionError'
