import datetime
import locale
from .util import Util

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import numpy
except ImportError:
    numpy = None


class Variable:
    """ 变量类 """

    OBJECT_PREFIX = '_@object_'
    REPEAT_PREFIX = '@repeat'

    def __init__(self, expr: str):
        self.sheet_name = None
        self.address = None
        self.address2 = None
        self.is_repeat = False
        self.has_at = False
        self.index = None
        self.expr = None
        if expr:
            self.load(expr)

    def clear(self):
        """ 清空 """
        self.sheet_name = None
        self.address = None
        self.address2 = None
        self.is_repeat = False
        self.has_at = False
        self.index = None
        self.expr = None

    def load(self, expr: str) -> 'Variable':
        """
        解析 expr 地址字符串, 返回 Variable 对象

        :param expr:  表达式, 可以是Excel地址，如： '{Sheet1!A1:B6｝', 也可以是一个Key名, 如：｛KeyName}
        :return:  如果是Excel地址，例如： address='{Sheet1!A1:B6｝', 返回tuple ('Sheet1', 'A1', 'B6').
                   如果不是Excel地址，例如： address='{KeyName}, 返回tuple (None, 'KeyName', None)
        """
        self.clear()

        if isinstance(expr, str):
            self.expr = expr
            expr = expr.strip()

            # 判断是否有大括号
            if Variable.is_var(expr):
                expr = expr[1:len(expr) - 1].strip()
                has_embrace = True
            else:
                raise ValueError('%s is not a variable, variable must start with "{" and end with "{"' % repr(expr))

            # 判断是否有 REPEAT_PREFIX
            if expr.startswith(Variable.REPEAT_PREFIX + ' '):
                self.is_repeat = True
                expr = expr[len(Variable.REPEAT_PREFIX):].strip()

            # 判断是否有 '@'
            if expr.startswith('@'):
                self.has_at = True
                expr = expr[1:].strip()

            # 判断是否有 '!'
            offset = expr.find('!')
            if offset > 0:
                self.sheet_name = expr[:offset].strip()
                expr = expr[offset + 1:].strip()
                # 判断是否有 ':'
                offset = expr.find(':')
                if offset > 0:
                    self.address2 = expr[offset + 1:]
                    self.address = expr[:offset]
                else:
                    self.address = expr
            elif has_embrace:
                self.address = expr

            # 取出 index
            if self.address.endswith(']') and self.address.find('[') > 0:
                index = self.address[self.address.find('[') + 1:]
                if index.endswith(']'):
                    index = index[:len(index)-1]
                self.index = index.strip()
                self.address = self.address[:self.address.find('[')].strip()

        return self

    @property
    def name(self):
        return self.address

    @staticmethod
    def is_var(word: str):
        """ 判断 word 是否是一个变量"""
        if isinstance(word, str) and word.startswith('{') and word.endswith('}'):
            return True
        return False


class VarData:
    SystemVars = ['index', 'indexA']
    """ 存有变量值的数据 """

    def __init__(self):
        self.path = ''

    @staticmethod
    def open(data):
        """ 打开数据 """
        if isinstance(data, dict):
            return DictVarData(data)
        elif Util.is_workbook(data):
            return ExcelVarData(data)
        elif type(data).__name__ == 'DataFrame' and hasattr(data, 'columns'):
            return DataFrameVarData(data)
        elif isinstance(data, str):
            Util.check_module(openpyxl, 'openpyxl')
            if Util.file_ext(data) == '.xlsx':
                obj = ExcelVarData(openpyxl.load_workbook(data, data_only=True))
                obj.path = Util.file_path(data)
                return obj

        raise ValueError("unknown variable data type %s" % repr(type(data)))

    # noinspection PyMethodMayBeStatic
    def _to_var(self, var):
        """ convert var to Variable object """
        if isinstance(var, Variable):
            return var
        elif isinstance(var, str):
            return Variable(var)
        raise ValueError("expect Variable but %s is found" % repr(type(var)))

    def has_var(self, var):
        """ 判断数据中是否有指定变量var的值 """
        var = self._to_var(var)
        if var.has_at:
            if var.address in VarData.SystemVars:
                return True

        raise NotImplementedError

    def var_length(self, var):
        """
        如果变量值是列表, 返回变量值列表长度.
        如果变量值不是列表，返回0
        """
        raise NotImplementedError

    def get_var(self, var, row=0):
        """
        取得有指定变量var的值

        :param var:  变量
        :param row:  (可选)当row=0, 返回第0行的值。当row>0, 返回第row行的值。
        :return: 返回对应行的变量值
        """
        var = self._to_var(var)

        # 处理特殊变量
        if var.has_at:
            if var.address == 'index':
                return str(row)
            elif var.address == 'indexA':
                return chr(ord('A') + row - 1)

        return self._get_var(var, row)

    def _get_var(self, var, row=0):
        raise NotImplementedError

    def close(self):
        pass


class ExcelVarData(VarData):
    """ 存有变量值的Excel数据 """

    def __init__(self, xlsx_obj):
        super().__init__()
        self.workbook = None
        if Util.is_workbook(xlsx_obj):
            self.workbook = xlsx_obj
            self.path = ''
        elif isinstance(xlsx_obj, str) and Util.file_ext(xlsx_obj) == '.xlsx':
            Util.check_module(openpyxl, 'openpyxl')
            self.workbook = openpyxl.load_workbook(xlsx_obj)
            self.path = Util.file_path(xlsx_obj)
        else:
            raise ValueError("expect Excel object but %s is found" % repr(type(xlsx_obj)))

    def has_sheet(self, sheet_name):
        """ 是否存在名为 sheet_name 的工作表 """
        if sheet_name:
            return sheet_name in self.workbook
        return False

    # noinspection PyMethodMayBeStatic
    def _convert_xls_date_format(self, fmt):
        """ 将 excel datetime 格式转换为 python 格式"""
        # TODO
        while fmt.find(']') > 0:
            fmt = fmt[fmt.find(']') + 1:]
        if fmt.endswith('@'):
            fmt = fmt[:len(fmt) - 1]
        if fmt.endswith(';'):
            fmt = fmt[:len(fmt) - 1]
        fmt = fmt.replace('yyyy', '%Y')
        fmt = fmt.replace('yy', '%y')
        fmt = fmt.replace('mm', '%m')
        fmt = fmt.replace('m', '%m')
        fmt = fmt.replace('dd', '%d')
        fmt = fmt.replace('d', '%d')
        fmt = fmt.replace('"', '')
        return fmt

    def has_var(self, var):
        """ 判断数据中是否有指定变量var的值 """
        var = self._to_var(var)
        if var.has_at:
            if var.address in VarData.SystemVars:
                return True

        if self.has_sheet(var.sheet_name):
            # noinspection PyBroadException
            try:
                sheet = self.workbook[var.sheet_name]
                cell = sheet[var.address]
                return cell.value is not None
            except Exception:
                return False

    def var_length(self, var):
        """
        如果变量值是列表, 返回变量值列表长度.
        如果变量值不是列表，返回0

        :param var: 变量
        :return: int
        """

        var = self._to_var(var)
        if self.has_sheet(var.sheet_name):
            # noinspection PyBroadException
            try:
                sheet = self.workbook[var.sheet_name]
                cell = sheet[var.address]
                # 统计单元格下方有数据的行数
                count = 0
                val = sheet.cell(cell.row + count + 1, cell.column).value
                while val is not None and val != '':
                    count += 1
                    val = sheet.cell(cell.row + count + 1, cell.column).value
                return count
            except Exception:
                pass
        return 0

    def _get_var(self, var, row=0):
        """
        取得有指定变量var的值

        :param var:  变量
        :param row:  (可选)当row=0, 返回第0行的值。当row>0, 返回第row行的值。
        :return: 返回对应行的变量值
        """
        var = self._to_var(var)
        if self.has_sheet(var.sheet_name):
            # noinspection PyBroadException
            try:
                sheet = self.workbook[var.sheet_name]
                cell = sheet[var.address]
                cell = sheet.cell(cell.row + row, cell.column)
                val = cell.value
                # process value
                if cell.is_date:
                    fmt = self._convert_xls_date_format(cell.number_format)
                    if isinstance(val, datetime.datetime):
                        if Util.has_chinese(fmt):
                            locale.setlocale(locale.LC_CTYPE, 'chinese')
                        val = val.strftime(fmt)
                elif isinstance(val, float):
                    val = round(val * 100) / 100  # TODO

                if var.has_at and isinstance(val, str):
                    val = Variable.OBJECT_PREFIX + val

                return val
            except Exception:
                pass

    def close(self):
        if self.path:
            self.workbook.close()
        self.path = ''


class DictVarData(VarData):
    """
    存有变量值的 dictionary 数据
    """
    def __init__(self, data):
        super().__init__()
        self.data = None
        if isinstance(data, dict):
            self.data = data
        else:
            raise ValueError("expect dict but %s is found" % repr(type(data)))

    def _get_value(self, key, default=None):
        """ get value from dictionary by key"""
        keys = key.split('.')
        d = self.data
        last_key = ''
        for k in keys:
            last_key = k
            if isinstance(d, dict) and k in d:
                d = d[k]
            elif isinstance(d, list) and Util.is_int(k) and len(d) > int(k):
                d = d[int(k)]
            else:
                return default, last_key
        return d, last_key

    def has_var(self, var):
        """
        判断数据中是否有指定变量var的值
        :param var:
        :return:
        """
        var = self._to_var(var)
        if var.has_at:
            if var.address in VarData.SystemVars:
                return True

        val, key = self._get_value(var.address)
        if val is None:
            return False

        if var.index is None:
            return val is not None
        else:
            # noinspection PyBroadException
            try:
                index = var.index
                if isinstance(val, list) or isinstance(val, tuple) or isinstance(val, str):
                    index = int(index)
                _ = val[index]
                return True
            except Exception:
                return False

    def var_length(self, var):
        """
        如果变量值是列表, 返回变量值列表长度.
        如果变量值不是列表，返回0

        :param var: 变量
        :return: int
        """
        var = self._to_var(var)
        val, key = self._get_value(var.address)
        return len(val) if isinstance(val, list) or isinstance(val, tuple) else 0

    def _get_var(self, var, row=0):
        """
        取得有指定变量var的值

        :param var:  变量
        :param row:  (可选)当row=0, 返回第0行的值。当row>0, 返回第row行的值。
        :return: 返回对应行的变量值
        """
        var = self._to_var(var)
        # 取得值和键名
        val, key = self._get_value(var.address)

        if val is None:
            return 0 if row < 0 else var.expr

        if row == 0:
            if isinstance(val, list):
                return key  # 当row=0， 返回键名
            else:
                if var.has_at and isinstance(val, str):
                    val = Variable.OBJECT_PREFIX + val
                return val
        else:
            if isinstance(val, list):
                # 当 row > 1, 返回列表项目
                val = val[row - 1] if len(val) >= row else None
                if var.has_at and isinstance(val, str):
                    val = Variable.OBJECT_PREFIX + val
                return val
            else:
                return None


class DataFrameVarData(VarData):
    """ 存有变量值的 DataFrame 数据 """
    def __init__(self, df):
        super().__init__()
        self.df = None
        if type(df).__name__ == 'DataFrame' and hasattr(df, 'columns'):
            self.df = df
        else:
            raise ValueError("expect DataFrame but %s is found" % repr(type(df)))

    # noinspection PyMethodMayBeStatic
    def _trim_quote(self, s):
        """ trim the leading and ending quote chars of the string """
        if isinstance(s, str):
            if s.startswith('"') and s.endswith('"'):
                return s[1: len(s)-1]
            elif s.startswith("'") and s.endswith("'"):
                return s[1: len(s) - 1]
        return s

    def has_var(self, var):
        """ 判断数据中是否有指定变量var的值 """
        var = self._to_var(var)
        if var.has_at:
            if var.address in VarData.SystemVars:
                return True

        # noinspection PyBroadException
        try:
            col_name = var.address
            if col_name == self.df.index.name:
                if var.index is None:
                    return True
                else:
                    index = int(var.index)
                    if index < len(self.df):
                        return True
            elif col_name in self.df.columns:
                if var.index is None:
                    return True
                else:
                    index = var.index
                    if self.df.index.dtype == numpy.int64 or self.df.index.dtype == numpy.int32:
                        index = int(index)
                    elif self.df.index.dtype == numpy.float32 or self.df.index.dtype == numpy.float64:
                        index = float(index)
                    else:
                        index = self._trim_quote(index)
                    _ = self.df.loc[index, col_name]
                    return True
        except Exception:
            pass

        return False

    def _get_var(self, var, row=0):
        """ 取得变量的值 """
        # noinspection PyBroadException
        try:
            var = self._to_var(var)
            col_name = var.address
            if col_name == self.df.index.name:
                # 如果 列名 是索引列
                if row == 0:
                    if var.index is None:
                        return col_name
                    else:
                        index = int(var.index)
                        return self.df.index[index]
                else:
                    return self.df.index[row - 1]
            elif col_name in self.df.columns:  # 如果 df 有这个列
                if row == 0:
                    if var.index is None:
                        return col_name  # 当row=0， 返回列名
                    else:
                        index = var.index
                        if self.df.index.dtype == numpy.int64 or self.df.index.dtype == numpy.int32:
                            index = int(index)
                        elif self.df.index.dtype == numpy.float32 or self.df.index.dtype == numpy.float64:
                            index = float(index)
                        else:
                            index = self._trim_quote(index)
                        return self.df.loc[index, col_name]
                else:
                    # 取得列的序号
                    col = self.df.columns.to_list().index(col_name)
                    if row - 1 < len(self.df):
                        return self.df.iloc[row - 1, col]
        except Exception:
            pass

    def var_length(self, var):
        """
        如果变量值是列表, 返回变量值列表长度.
        如果变量值不是列表，返回0
        """
        var = self._to_var(var)
        # 当 data 是 DataFrame
        col_name = var.address
        if col_name == self.df.index.name:
            return len(self.df)
        elif col_name in self.df.columns:
            return len(self.df)
        return 0
