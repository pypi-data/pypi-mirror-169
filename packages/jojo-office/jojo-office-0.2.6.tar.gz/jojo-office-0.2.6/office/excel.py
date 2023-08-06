#
# Excel : process the data in .xlsx file
#
# Example:
# ::
#     import office
#
#     # open file datafile.xlsx, return workbook object
#     workbook = office.open_file("datafile.xlsx")
#
#     # list worksheets
#     sheet_names = workbook.sheetnames
#     print(sheet_names)
#
#     # get a sheet
#     sheet = workbook['Sales']
#     print(sheet.name)
#
#     # get a cell
#     cell = sheet['B2']
#     print(cell.value)
#
#     # write a cell
#     cell.value = 'New City'
#
#     # save
#     workbook.save("newfile.xlsx")
#
#     # close
#     workbook.close()
#
#

import os
import time
import inspect
from .base import BaseFile
from .util import Util
from .win32 import Win32
import re
import openpyxl

from openpyxl.styles.numbers import FORMAT_GENERAL, FORMAT_TEXT, FORMAT_NUMBER
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, FORMAT_DATE_DATETIME
from openpyxl.styles import PatternFill, GradientFill

try:
    import xlwings
except ImportError:
    xlwings = None


try:
    import pandas
    from pandas.io.formats.excel import ExcelFormatter
    from pandas import ExcelWriter
except ImportError:
    pandas = None
    ExcelFormatter = None
    ExcelWriter = None


class XlsConsts:
    # https://docs.microsoft.com/en-us/office/vba/api/word.wdsaveformat
    wdFormatDocument = 0   # Microsoft Office Word 97 - 2003 binary file format.
    wdFormatTemplate = 1  # Word template format.
    wdFormatText = 2  # Microsoft Windows text format.
    wdFormatRTF = 6  # Rich text format (RTF).
    wdFormatHTML = 8  # Standard HTML format.
    wdFormatWebArchive = 9  # Web archive format.
    wdFormatFilteredHTML = 10  # Filtered HTML format.
    wdFormatXMLTemplate = 14  # XML template format.
    wdFormatPDF = 17  # PDF format.


class XlsUtil:

    @staticmethod
    def is_formula(value):
        """ whether value is a formula of Excel"""
        if isinstance(value, str) and value.startswith('='):
            return True
        return False

    @staticmethod
    def is_hex_char(c: str):
        """ whether char c is a hexadecimal char """
        n = ord(c)
        if (ord('0') <= n <= ord('9')) or (ord('A') <= n <= ord('F')):
            return True
        return False

    @staticmethod
    def is_hex_str(s: str):
        """ whether the string is a hexadecimal string """
        if not isinstance(s, str):
            return False

        s = s[1:] if s.startswith('#') else s
        s = s.upper()
        for c in s:
            if not XlsUtil.is_hex_char(c):
                return False
        return True

    @staticmethod
    def to_hex_str(n: int):
        """ convert specified number to hexadecimal string """
        s = hex(n)
        if s.startswith('0x'):
            s = s[2:]
        if len(s) % 2 == 1:
            s += '0'
        return s.upper()

    @staticmethod
    def is_int_str(s: str):
        """ whether the string is an integer string"""
        if not isinstance(s, str):
            return False

        if s == '':
            return False

        for c in s:
            if not ('0' <= c <= '9'):
                return False
        return True

    @staticmethod
    def is_color(value):
        """ whether the value is a color string """
        if isinstance(value, str) and XlsUtil.is_hex_str(value) and len(value) >= 3:
            return True
        return False

    @staticmethod
    def to_color_str(value):
        """  convert the value to a color string"""
        if isinstance(value, str) and XlsUtil.is_hex_str(value):
            if value.startswith('#'):
                value = value[1:]
                if len(value) == 3:
                    value = value[0] + value[0] + value[1] + value[1] + value[2] + value[2]
            else:
                value = value.upper()

            if value.startswith('00') and len(value) == 8:
                value = value[2:]
        return value

    @staticmethod
    def is_number(value):
        """ whether the value is number """
        if isinstance(value, int) or isinstance(value, float):
            if not isinstance(value, bool):
                return True
        return False

    @staticmethod
    def rgb(r: int, g: int, b: int) -> str:
        """ return rgb color string """
        return XlsUtil.to_hex_str(r) + XlsUtil.to_hex_str(g) + XlsUtil.to_hex_str(b)

    @staticmethod
    def convert_number_format(value):
        """ convert number format to openpyxl format """
        if value is None or value == '':
            return FORMAT_GENERAL

        if value == 'text' or value == '%s':
            return FORMAT_TEXT

        if value == 'number':
            return FORMAT_NUMBER

        if value == 'percentage' or value == 'percent':
            return FORMAT_PERCENTAGE

        if value == 'date' or value == 'datetime':
            return FORMAT_DATE_DATETIME

        if re.search('%[YyHhMmS]', value) is not None:
            value = value.replace('%Y', 'yyyy')
            value = value.replace('%y', 'yy')
            value = value.replace('%m', 'mm')
            value = value.replace('%d', 'dd')
            value = value.replace('%H', 'hh')
            value = value.replace('%M', 'mm')
            value = value.replace('%S', 'ss')
            return value

        if re.search('%(.*)f', value) is not None:
            value = XlsUtil.to_excel_number_format(value)

        return value

    @staticmethod
    def to_excel_number_format(format_str: str):
        """
        convert printf() float format string to Excel number format

        :param format_str:  float format string in printf(), such as '%8.3f'

        :return: return Excel number format, example:  '%8.3f'  ->  #,###.000
        """
        value = format_str

        # find substring like '%??f'
        m = re.match('%(.*)f', value)
        if m is None:
            return '0'

        # get the '%??f' substring
        span = m.span()
        s = value[span[0] + 1: span[1] - 1]

        # split by char '.'
        arr = s.split('.')
        suffix = value[span[1]:].strip()

        if len(arr) == 1 and arr[0] != '':
            length = int(arr[0]) if XlsUtil.is_int_str(arr[0]) else 1
            length = -length if length < 0 else length

            if suffix == ',':
                return '#,###'
            else:
                return '#' * length

        elif len(arr) == 2:
            length = int(arr[0]) if XlsUtil.is_int_str(arr[0]) else 1
            length = -length if length < 0 else length

            digit = int(arr[1]) if XlsUtil.is_int_str(arr[1]) else 0
            length = digit + 1 if length < digit else length

            if suffix == ',':
                return '#,###' + '.' + ('0' * digit)
            else:
                return ('#' * (length - digit)) + '.' + ('0' * digit)
        else:
            return '0'

    @staticmethod
    def col_str_to_int(col_str: str):
        """
        convert Excel column string to integer, such as 'A' -> 1,  'D' -> 4

        :param col_str:  column string, such as 'A', 'BZ'

        :return: return int.  example: 'A' -> 1,  'D' -> 4
        """
        col_str = col_str.upper()
        num = 0
        for i in range(0, len(col_str)):
            num *= 26
            c = col_str[i]
            if ord('A') <= ord(c) <= ord('Z'):
                num += ord(c) - ord('A') + 1
            else:
                return -1
        return num

    @staticmethod
    def col_int_to_str(col_num: int):
        """
        convert integer to Excel column string. such as 1 -> 'A',  4 -> 'D'

        :param col_num: column number ( First column 'A' is 1)

        :return: return string, example: 1 -> 'A',  4 -> 'D'
        """
        result = ''
        while True:
            if col_num > 26:
                col_num, r = divmod(col_num - 1, 26)
                result = chr(r + ord('A')) + result
            else:
                return chr(col_num + ord('A') - 1) + result

    @staticmethod
    def is_list_like(obj):
        """
        Whether the object is like a list

        :param obj: the object

        :return:
        """
        if isinstance(obj, list) or isinstance(obj, tuple):
            return True

        if (not isinstance(obj, str)) and hasattr(obj, '__getitem__') \
                and hasattr(obj, '__len__') and (not isinstance(obj, dict)):
            return True

        return False


class Cell:
    def __init__(self, sheet, obj):
        self.sheet: WorkSheet = sheet
        self.cell_obj = None
        self._load(obj)

    def _load(self, obj):
        """
        read cell values

        :param obj: Address string or tuple(row, col) or cell object
        :return:
        """
        if isinstance(obj, str):
            c = self.sheet[obj]
            if isinstance(c, Cell):
                self.cell_obj = c.cell_obj
            elif isinstance(c, Range):
                self.cell_obj = c.cell1.cell_obj
        elif Util.is_list_like(obj) and len(obj) == 2 and \
                isinstance(obj[0], int) and isinstance(obj[1], int):
            self.cell_obj = self.sheet.cell(obj[0], obj[1]).cell_obj
        else:
            self.cell_obj = obj

    @property
    def value(self):
        """ value of the Cell """
        return self.cell_obj.value

    @value.setter
    def value(self, val):
        if type(val).__name__ == 'DataFrame':
            self.sheet.write_dataframe(self.address, val)
        else:
            self.cell_obj.value = val

    @property
    def row(self):
        """ row of the Cell, first row is 1 """
        return self.cell_obj.row

    @property
    def column(self):
        """ column of the cell, column 'A' is 1 """
        return self.cell_obj.column

    @property
    def address(self):
        """ address of the Cell, such as 'A3' """
        return self.to_str(absolute=False)

    def to_str(self, absolute=False):
        """
        convert to string.

        :param absolute: (optional) whether display absolute mark '$'
        :return:  return a string
        """
        result = ''
        if self.column >= 0:
            if absolute:
                result += '$'
            result += XlsUtil.col_int_to_str(self.column)

        if self.row >= 0:
            if absolute:
                result += '$'
            result += str(self.row)

        return result

    def move(self, delta_row: int, delta_col: int):
        """
        Move the cell to a new place.

        :Chinese: 将单元格移动到一个新的位置.

        :param delta_row:   the rows to move. negative number means move left
        :param delta_col:  the columns to move. negative number means move up
        :return:  self
        """
        new_row = self.row + delta_row
        new_col = self.column + delta_col
        if new_row <= 0 or new_col <= 0:
            raise ValueError("row or column is little than 0")
        self.cell_obj = self.sheet.cell(new_row, new_col).cell_obj
        return self

    def expand(self, right=0, down=0):
        """
        expand the cell to a range.

        :Chinese: 水平或垂直方向 扩展区域

        :param right:   the rows to expand rightward. 0 means expand to non-empty cell, None means not expand
                :Chinese: 水平方向右扩展区域的列数，0表示向右扩展到非空单元格，None表示不扩展。

        :param down:   the columns to downward. 0 means expand to non-empty cell, None means not expand
                :Chinese: 垂直方向向下扩展区域的行数，0表示向下扩展到非空单元格，None表示不扩展。

        :return:  Range
        """
        return Range(self.sheet, self).expand(right, down)  # TODO

    def equals(self, cell):
        """ whether current cell equals to specified cell """
        if self.sheet.sheet_obj == self.sheet.sheet_obj:
            if self.row == cell.row and self.column == cell.column:
                return True

    @property
    def column_width(self):
        """ width of the column """
        if self.sheet.book.is_xlwings:
            return self.cell_obj.column_width
        else:
            col = XlsUtil.col_int_to_str(self.column)
            return self.sheet.sheet_obj.column_dimensions[col].width

    @column_width.setter
    def column_width(self, value):
        if self.sheet.book.is_xlwings:
            self.cell_obj.column_width = value
        else:
            col = XlsUtil.col_int_to_str(self.column)
            self.sheet.sheet_obj.column_dimensions[col].width = value

    @property
    def row_height(self):
        """ height of the row """
        if self.sheet.book.is_xlwings:
            return self.cell_obj.row_height
        else:
            return self.sheet.sheet_obj.row_dimensions[self.row].height

    @row_height.setter
    def row_height(self, value):
        if self.sheet.book.is_xlwings:
            self.cell_obj.row_height = value
        else:
            self.sheet.sheet_obj.row_dimensions[self.row].height = value

    @property
    def number_format(self):
        """ number format """
        return self.cell_obj.number_format

    @number_format.setter
    def number_format(self, value):
        self.cell_obj.number_format = XlsUtil.convert_number_format(value)

    @property
    def color(self):
        """
        Gets and sets the background color of the specified Range.

        To set the color, either use an RGB tuple ``(0, 0, 0)`` or a hex string
        like ``#efefef`` or an Excel color constant.

        :return:
        """
        if self.sheet.book.is_xlwings:
            return self.cell_obj.color
        else:
            fill = self.cell_obj.fill
            if fill.patternType == "solid":
                return XlsUtil.to_color_str(fill.start_color.rgb)
            else:
                return None

    @color.setter
    def color(self, value):
        """
        Gets and sets the background color of the specified Range.

        To set the color, either use an RGB tuple ``(0, 0, 0)`` or a hex string
        like ``#efefef`` or an Excel color constant.
        """
        if self.sheet.book.is_xlwings:
            self.cell_obj.color = value
        else:
            if value is None:
                self.cell_obj.fill = PatternFill("none")
            else:
                self.cell_obj.fill = PatternFill("solid", start_color=XlsUtil.to_color_str(value))

    def __repr__(self):
        return '<Cell %s>' % repr(self.address)


class RangeList(list):
    pass


class Range:
    def __init__(self, sheet, obj, cell2=None):
        self.sheet: WorkSheet = sheet
        self.cell1 = None
        self.cell2 = None
        self.range_obj = None
        self._load(obj, cell2)

    def _load(self, obj, cell2=None):
        """
        设置Range区域

        :param obj:  可以是一个地址字符串，可以是Cell对象，可以是 xlwings的Range对象， openpyxl的tuple
        :param cell2:  可以是Cell对象，可以是 xlwings的Range对象， openpyxl的tuple
        :return: None
        """
        if isinstance(obj, str):
            obj = self.sheet[obj]

        if self.sheet.book.is_xlwings:
            if type(obj).__name__ == 'Range':
                self.range_obj = obj
                self.cell1 = self.sheet.cell(self.range_obj.row, self.range_obj.column)
                self.cell2 = self.sheet.cell(self.range_obj.row + len(self.range_obj.rows) - 1,
                                             self.range_obj.column + len(self.range_obj.columns) - 1)
            elif isinstance(obj, Cell):
                self.cell1 = obj
                self.cell2 = cell2 if isinstance(cell2, Cell) else self.cell1
                self.rect()
                self.range_obj = self.sheet.sheet_obj[self.cell1.address + ':' + self.cell2.address]
            elif Util.is_list_like(obj) and len(obj) == 2 and Util.is_list_like(cell2) and len(cell2) == 2:
                self.cell1 = self.sheet.cell(obj[0], obj[1])
                self.cell2 = self.sheet.cell(cell2[0], cell2[1])
                self.rect()
                self.range_obj = self.sheet.sheet_obj[self.cell1.address + ':' + self.cell2.address]
            else:
                raise ValueError("%s is not a Range" % repr(type(obj)))
        else:
            if isinstance(obj, tuple):
                row = obj[0]
                self.cell1 = Cell(self, row[0])

                row = obj[len(obj) - 1]
                self.cell2 = Cell(self, row[len(row) - 1])
                self.rect()
            elif isinstance(obj, Cell):
                self.cell1 = obj
                self.cell2 = cell2 if isinstance(cell2, Cell) else self.cell1
                self.rect()
                self.range_obj = None
            else:
                raise ValueError("%s is not a Range" % repr(type(obj)))

    def rect(self):
        """ return start_row, start_col, end_row, end_col of the rectangle of the Range """
        start_row = self.cell1.row
        start_col = self.cell1.column
        end_row = self.cell2.row
        end_col = self.cell2.column
        if end_row < start_row:
            start_row, end_row = end_row, start_row
        if end_col < start_col:
            start_col, end_col = end_col, start_col
        if self.cell1.row != start_row or self.cell1.column != start_col:
            self.cell1 = self.sheet.cell(start_row, start_col)
        if self.cell2.row != end_row or self.cell2.column != end_col:
            self.cell2 = self.sheet.cell(end_row, end_col)
        return start_row, start_col, end_row, end_col

    @property
    def value(self):
        """ value of the range, return a list """
        if self.sheet.book.is_xlwings:
            return self.range_obj.value

        else:
            result = []
            for r in range(self.cell1.row, self.cell2.row + 1):
                data = []
                for c in range(self.cell1.column, self.cell2.column + 1):
                    data.append(self.sheet.cell(r, c).value)
                result.append(data)
            return result

    @value.setter
    def value(self, val):
        if self.sheet.book.is_xlwings:
            self.range_obj.value = val

        else:
            if not Util.is_list_like(val):
                val = [[val]]

            if Util.is_table_list(val):
                start_row, start_col, end_row, end_col = self.rect()
                for r in range(0, end_row - start_row + 1):
                    for c in range(0, end_col - start_col + 1):
                        row_val = val[r] if len(val) > r else []
                        if len(row_val) > c:
                            self.sheet.cell(start_row + r, start_col + c).value = row_val[c]

    def clear(self, value=None):
        """
        Clear the content of the range, fill each cell with value.

        :Chinese: 清空区域， 填入默认值 value

        :param value: (optional) value which is filled into cells
        :return: self
        """
        # scan the area, set each cell's value
        for row in range(self.cell1.row, self.cell2.row + 1):
            for col in range(self.cell1.column, self.cell2.column + 1):
                self.sheet.cell(row, col).value = value
        return self

    @property
    def address(self):
        """ address of the Range, such as 'A3:B6' """
        return self.cell1.address + ':' + self.cell2.address

    @property
    def rows(self):
        """ Return a list the rows in the range """
        rs = RangeList()
        start_row, start_col, end_row, end_col = self.rect()
        for row in range(start_row, end_row + 1):
            cell1 = self.sheet.cell(row, start_col)
            cell2 = self.sheet.cell(row, end_col)
            rs.append(Range(self.sheet, cell1, cell2))
        return rs

    @property
    def columns(self):
        """ Return a list the columns in the range """
        rs = RangeList()
        start_row, start_col, end_row, end_col = self.rect()
        for col in range(start_col, end_col + 1):
            cell1 = self.sheet.cell(start_row, col)
            cell2 = self.sheet.cell(end_row, col)
            rs.append(Range(self.sheet, cell1, cell2))
        return rs

    def expand(self, right=0, down=0):
        """
        expand the range.

        :Chinese: 水平或垂直方向 扩展区域

        :param right:   the rows to expand rightward. 0 means expand to non-empty cell, None means not expand
                :Chinese: 水平方向右扩展区域的列数，0表示向右扩展到非空单元格，None表示不扩展。

        :param down:   the columns to downward. 0 means expand to non-empty cell, None means not expand
                :Chinese: 垂直方向向下扩展区域的行数，0表示向下扩展到非空单元格，None表示不扩展。

        :return:  self
        """
        row = self.cell2.row
        col = self.cell2.column
        start_col = col

        # expand horizontal
        if isinstance(right, int):
            if right == 0:
                cell = self.sheet.cell(row, col + 1)
                while cell.value is not None:
                    col += 1
                    cell = self.sheet.cell(row, col + 1)
            else:
                col += right
                if col < 1:
                    col = 1

        # expand vertical
        if isinstance(down, int):
            if down == 0:
                cell = self.sheet.cell(row + 1, start_col)
                while cell.value is not None:
                    row += 1
                    cell = self.sheet.cell(row + 1, start_col)
            else:
                row += down
                if row < 1:
                    row = 1

        self.cell2 = self.sheet.cell(row, col)
        return self

    def to_list(self):
        """ read values of the cells in the range, return a list """
        return self.value

    def to_dataframe(self):
        """ read values of the cells in the range, return a DataFrame """
        return self.sheet.read_dataframe(self)

    def merge_cells(self):
        """ merge cells in the range """
        if self.sheet.book.is_xlwings:
            self.range_obj.merge_cells()
        else:
            self.sheet.sheet_obj.merge_cells(self.address)

    def is_cell(self):
        """ whether the range contains only one cell """
        if self.cell1.row == self.cell2.row and self.cell1.column == self.cell2.column:
            return True
        return False

    def __repr__(self):
        return '<Range %s>' % repr(self.address)

    @property
    def color(self):
        """ Get/Set the background color of the cells in the range"""
        return self.cell1.color

    @color.setter
    def color(self, value):
        # scan the area, set each cell's value
        for row in range(self.cell1.row, self.cell2.row + 1):
            for col in range(self.cell1.column, self.cell2.column + 1):
                self.sheet.cell(row, col).color = value


class WorkSheet:
    def __init__(self, book, sheet_obj):
        self.book: Excel = book
        self.sheet_obj = sheet_obj

    @property
    def max_column(self):
        """ the max columns that worksheet used  """
        if self.book.is_xlwings:
            all_cells = self.sheet_obj.cells
            count = all_cells.columns.count
            max_c = 1
            for row in range(1, 200):
                r = 100 + (row - 100) * 2 if row > 100 else row
                cur_c = self.range((r, count)).range_obj.end('left').column
                max_c = max(max_c, cur_c)
            return max_c + 1
        else:
            return self.sheet_obj.max_column

    @property
    def max_row(self):
        """ the max row that worksheet used """
        if self.book.is_xlwings:
            all_cells = self.sheet_obj.cells
            count = all_cells.rows.count
            max_r = 1
            for col in range(1, 200):
                c = 100 + (col - 100) * 2 if col > 100 else col
                cur_r = self.range((count, c)).range_obj.end('up').row
                max_r = max(max_r, cur_r)
            return max_r + 1
        else:
            return self.sheet_obj.max_row

    @property
    def title(self):
        """ the title of the worksheet ( alias of name) """
        if self.book.is_xlwings:
            return self.sheet_obj.name
        else:
            return self.sheet_obj.title

    @title.setter
    def title(self, value):
        if self.book.is_xlwings:
            self.sheet_obj.name = value
        else:
            self.sheet_obj.title = value

    @property
    def name(self):
        """ the name of the sheet """
        return self.title

    @name.setter
    def name(self, value):
        self.title = value

    def cell(self, row: int, column: int):
        """
        Get a cell of specified row, column

        :Chinese: 根据行号、列号，返回一个Cell对象

        :param row:   row number(int), the first row is 1
        :param column: column number(int), the first column is 1
        :return: return a Cell objecdt
        """
        if self.book.is_xlwings:
            return Cell(self, self.sheet_obj.range((row, column)))
        else:
            return Cell(self, self.sheet_obj.cell(row, column))

    def range(self, cell1, cell2=None):
        """
        get a range

        :param cell1: left-top cell
        :param cell2: (optinal)right-bottom cell
        :return: return a Range object
        """
        cell2 = cell1 if cell2 is None else cell2
        return Range(self, cell1, cell2)

    def __getitem__(self, item) -> [Cell, Range]:
        """ get a cell or a range """

        if isinstance(item, str):
            obj = self.sheet_obj[item]
            if self.book.is_xlwings:
                if len(obj.rows) == 1 and len(obj.columns) == 1:
                    return Cell(self, obj)
                else:
                    return Range(self, obj)
            else:
                if isinstance(obj, tuple):
                    return Range(self, obj)
                else:
                    return Cell(self, obj)

    @property
    def index(self):
        """ return the index of the sheet in the parent workbook """
        if self.book.is_xlwings:
            return self.sheet_obj.index
        else:
            return self.title in self.sheet_obj.sheetnames

    def find(self, value1, value2=None, regex=False):
        """
        Find a cell

        :param value1:   first value
        :param value2:  (optional) second value
        :param regex:   (optional) whether use regex

        :return: return Cell object if found. return None if not found.
        """
        if value2 is not None:
            c1 = self.find(value1, regex=regex)
            c2 = self.find(value2, regex=regex)
            if c1 and c2:
                return Cell(self, (c1.row, c2.column))

        max_row = self.max_row
        max_column = self.max_column
        for row in range(1, max_row):
            for col in range(1, max_column):
                val = self.cell(row, col).value
                if not regex:
                    if value1 == val:
                        return Cell(self, (row, col))
                else:
                    if isinstance(val, str) and re.search(value1, str(val)) is not None:
                        return Cell(self, (row, col))

        return None

    def _address_to_range(self, address):
        """ convert address to Range """
        if isinstance(address, Range):
            return address
        else:
            c = self[address]
            if isinstance(c, Cell):
                return Range(c.sheet, c, c)
            else:
                return c

    def read_list(self, address):
        """
        read list from specified address

        :param address: the address, such as: 'A2:C4'
        :return: return a list
        """
        rng = self._address_to_range(address)
        return rng.to_list()

    def write_list(self, address, list_data, vertical=False):
        """
        write a list to the specified address

        :param address:   the address, such as: 'A2:C4'
        :param list_data: list of data
        :param vertical:  (optional)whether write vertically
        :return:
        """
        rng = self._address_to_range(address)
        start_row, start_col, end_row, end_col = rng.rect()

        for r, row_data in enumerate(list_data):
            if Util.is_list_like(row_data):
                for c, val in enumerate(row_data):
                    if not vertical:
                        row = start_row + r
                        col = start_col + c
                    else:
                        row = start_row + c
                        col = start_col + r
                    if (row <= end_row or end_row == 0) and (col <= end_col or end_col == 0):
                        self.cell(row, col).value = val
            else:
                if not vertical:
                    row = start_row
                    col = start_col + r
                else:
                    row = start_row + r
                    col = start_col
                self.cell(row, col).value = row_data

    def read_dataframe(self, address):
        """
         read DataFrame from specified address

        :param address:  the address, such as: 'A2:C4'
        :return: DataFrame
        """
        data = self.read_list(address)
        if len(data) > 0:
            return pandas.DataFrame(data[1:], columns=data[0])
        else:
            return pandas.DataFrame()

    def write_dataframe(self, address, df, header=True, vertical=False, **kwargs):
        """
        write pandas DataFrame

        :param address:     the address
        :param df:          DataFrame object
        :param header:      whether write header of the DataFrame object
        :param vertical:    vertical placed
        :return: return a tuple of (rows, cols) which are the count of rows and cols that be written.
        """
        if not self.book.is_xlwings and not vertical:
            self._openpyxl_write_dataframe(address, df, header, **kwargs)
            return

        rng = self._address_to_range(address)
        start_row, start_col, end_row, end_col = rng.rect()

        h = 1 if header else 0

        if end_row == start_row and end_col == start_col:
            if vertical:
                end_row = start_row + len(df.columns) - 1 if len(df.columns) > 0 else start_row
            else:
                end_row = start_row + len(df) - 1 + h if len(df) > 0 else start_row

        if end_col == start_col and end_col == start_col:
            if vertical:
                end_col = start_col + len(df) - 1 + h if len(df) > 0 else end_col
            else:
                end_col = start_col + len(df.columns) - 1 if len(df.columns) > 0 else end_col

        count = 0
        # write dataframe columns
        if header:
            if vertical:
                r = Range(self, (start_row, start_col + count), (end_row, start_col + count))
            else:
                r = Range(self, (start_row + count, start_col), (start_row + count, end_col))
            self.write_list(r, list(df.columns), vertical)
            count += 1

        # write dataframe data of each row
        for i in range(len(df)):
            if vertical:
                r = Range(self, (start_row, start_col + count), (end_row, start_col + count))
            else:
                r = Range(self, (start_row + count, start_col), (start_row + count, end_col))
            self.write_list(r, list(df.iloc[i]), vertical)
            count += 1

    def _openpyxl_write_dataframe(self, address, df, header=True, **kwargs):
        rng = self[address]
        if isinstance(rng, Cell):
            start_cell = rng
        else:
            start_cell = rng.cell1

        sheet_name = self.name
        writer = ExcelWriter(self.book.filename, engine='openpyxl')
        writer.book = self.sheet_obj
        writer._get_sheet = dict((ws.title, ws) for ws in self.book.book_obj.worksheets)

        index = kwargs.pop('index') if 'index' in kwargs else False
        header = kwargs.pop('header') if 'header' in kwargs else header

        ExcelFormatter.header_style = None

        self.book.save()
        df.to_excel(writer, sheet_name=sheet_name, index=index, header=header,
                    startrow=start_cell.row - 1, startcol=start_cell.column - 1, **kwargs)

        writer.save()
        writer.close()

        delta = 1 if header else 0
        return len(df) + delta, len(df.columns)

    def protect(self, password: str):
        """ Protect the Worksheet with password """
        if self.book.is_xlwings:
            self.sheet_obj.api.Protect(password)
        else:
            self.sheet_obj.protection.sheet = True
            self.sheet_obj.protection.password = password
            self.sheet_obj.protection.enable()

    def unprotect(self, password: str):
        """ Unprotect the WorkSheet """
        if self.book.is_xlwings:
            self.sheet_obj.api.Unprotect(password)
        else:
            self.sheet_obj.protection.password = password
            self.sheet_obj.protection.disable()
            self.sheet_obj.protection.sheet = False

    def move(self, offset):
        """
         Move the sheet

        :param offset: distance to move
        :return: self
        """
        self.book.move_sheet(self, offset)

    def delete(self):
        """ delete the worksheet """
        self.book.delete_sheet(self)

    def activate(self):
        """ activate the worksheet """
        self.book.active = self

    def __repr__(self):
        return '<WorkSheet %s>' % repr(self.name)


class WorkSheets:
    def __init__(self, book):
        self.book = book

    def __contains__(self, sheet_name):
        return self.book.__contains__(sheet_name)

    def __getitem__(self, item) -> WorkSheet:
        return self.book[item]

    @property
    def count(self):
        """ count of worksheets """
        return self.book.count

    @property
    def active(self) -> WorkSheet:
        return self.book.active

    @active.setter
    def active(self, value):
        self.book.active = value

    def add(self, name=None, index=None):
        """ add a sheet """
        self.book.create_sheet(name, index)

    def delete(self, name=None):
        """ delete a worksheet """
        self.book.delete_sheet(name)


# Excel .xlsx document
class Excel(BaseFile):
    def __init__(self, filename=None, template=None, **kwargs):
        """
        init and open file

        :param filename:  the filename to open.
                        if filename is None, create new empty file.
                        if filename is None and run in xlwings, get the workbook that calls,
                        using xlwings.Book.call()

        :param template: (optinal) template filename. If not None, use the template
                        file to create new file.
        :param kwargs:
        """
        self.is_xlwings = False
        self.xlwings_book = None
        self.msg = ''
        self._file_password = None

        # noinspection PyBroadException
        try:
            self.xlwings_book = xlwings.Book.caller()
        except Exception:
            # Exception: Book.caller() must not be called directly. Call through Excel
            pass

        if not self.xlwings_book and filename is None:
            caller_file = inspect.stack()[1].filename
            same_name_xlsx = Util.change_file_ext(caller_file, '.xlsx')
            if os.path.exists(same_name_xlsx):
                filename = same_name_xlsx

        super().__init__(filename, template, **kwargs)

    @property
    def book_obj(self):
        """ Workbook object of openpyxl or book object of xlwings """
        return self._obj

    def _new_file(self, filename, **kwargs):
        if self.xlwings_book:
            self._obj = self.xlwings_book
            self.filename = self.xlwings_book.fullname
            self.is_xlwings = True
            self.msg = '_new_file ' + str(filename)
        else:
            self._obj = openpyxl.Workbook()
            self.is_xlwings = False
            self.filename = None
            # modify 'Sheet' title, make it looks like Microsoft Excel new file
            if 'Sheet' in self._obj:
                self._obj['Sheet'].title = 'Sheet1'
            if filename:
                self.save(filename, **kwargs)
            return self

    def _open(self, filename, **kwargs):
        read_only = kwargs.get('read_only', None)
        data_only = kwargs.get('data_only', None)
        if self.xlwings_book is not None:
            self._obj = self.xlwings_book
            self.is_xlwings = True
        else:
            self._obj = openpyxl.load_workbook(filename, read_only=read_only, data_only=data_only)
            self.is_xlwings = False
            self.filename = filename
        return True

    def _close(self, **kwargs):
        if self._obj is not None:
            save = kwargs.get('save', False)

            # save before close
            if save and self.filename != '':
                self.save()

            # noinspection PyBroadException
            try:
                self._obj.close()
            finally:
                self._obj = None
                self.filename = ''
        return True

    def _save_file(self, filename, **kwargs):
        formats = {
            '.rtf': XlsConsts.wdFormatRTF,
        }

        file_ext = Util.file_ext(filename)

        if file_ext == '.xlsx':
            return self._save_xlsx(filename, **kwargs)

        elif file_ext in formats:
            app, workbook = Win32.open_excel(self.filename, False)
            if not Util.is_absolute_path(filename):
                filename = os.path.join(os.getcwd(), filename)
            workbook.SaveAs(filename, formats[file_ext], **kwargs)
            Win32.close(app, workbook)

        else:
            return super()._save_file(filename, **kwargs)

    def _save_xlsx(self, filename, **kwargs):
        if self._obj is not None:
            filename = self.filename if filename is None or filename == '' else filename
            self._obj.save(filename, **kwargs)
            self.filename = filename

    def __contains__(self, name):
        """ whether worksheet of specified name exists """
        if isinstance(name, WorkSheet):
            name = name.title

        if self.is_xlwings:
            return name in self.sheetnames
        else:
            return name in self.book_obj

    @property
    def sheets(self):
        """ return a collection of worksheets """
        return WorkSheets(self)

    @property
    def sheetnames(self):
        """ return a list of name of worksheets """
        if self.is_xlwings:
            return [sheet.name for sheet in self.book_obj.sheets]
        else:
            return self.book_obj.sheetnames

    @property
    def count(self):
        """ return count of worksheet """
        if self.is_xlwings:
            return self.book_obj.sheets.count
        else:
            return len(self.book_obj.sheetnames)

    def _get_sheet_by_name(self, name: str) -> WorkSheet:
        """ get sheet by name """
        if name in self:
            if self.is_xlwings:
                return WorkSheet(self, self.book_obj.sheets[name])
            else:
                return WorkSheet(self, self.book_obj[name])

    def _get_sheet_by_index(self, index: int) -> WorkSheet:
        """ get sheet object by index"""
        if 0 <= index < self.count:
            if self.is_xlwings:
                return WorkSheet(self, self.book_obj.sheets[index])
            else:
                return WorkSheet(self, self.book_obj.worksheets[index])

    def _get_sheet(self, item, create=True) -> WorkSheet:
        """
        Get worksheet object

        :Chinese: 取得 WorkSheet 对象。

        :param item: （(Optional) tile of sheet, or index of sheet
        :return: return WorkSheet object
        """
        if isinstance(item, WorkSheet):
            return item

        sheet = None

        if isinstance(item, str):
            if item not in self and create:
                sheet = self.create_sheet(item)
            else:
                sheet = self._get_sheet_by_name(item)

        elif isinstance(item, int):
            if item >= self.count and create:
                for n in range(self.count, item + 1):
                    name = 'Sheet' + str(n)
                    count = 1
                    while name in self:
                        name = name + '(' + str(count) + ')'
                        count += 1
                    self.create_sheet(name)
            # find sheet
            sheet = self._get_sheet_by_index(item)

        # return
        if sheet:
            return sheet
        else:
            raise ValueError('WorkSheet %s not exists' % repr(item))

    def create_sheet(self, name=None, index=None):
        """
        Create a new WorkSheet.

        :param name:  name of new Sheet
        :param index:  (optional int, index of the worksheet.
            Default if None.(the worksheet will be put at then end)
        :return: self
        """
        if self.is_xlwings:
            sheet_count = self.book_obj.sheets.count
            before = after = None
            if sheet_count > 0:
                if index is None or 0 > index >= sheet_count:
                    after = self.book_obj.sheets[sheet_count - 1]
                elif index == 0:
                    before = self.book_obj.sheets[0]
                else:
                    before = self.book_obj.sheets[index]
            ws = self.book_obj.sheets.add(name, before=before, after=after)
            return WorkSheet(self, ws)
        else:
            ws = self.book_obj.create_sheet(title=name, index=index)
            return WorkSheet(self, ws)

    def move_sheet(self, sheet, offset=0):
        """
        Move the sheet to a new place.

        :Chinese: 移动WorkSheet

        :param sheet:  worksheet, could be name or WorkSheet object
        :param offset: distance to move
        :return: self
        """
        if self.count == 0:
            return

        sheet = self._get_sheet(sheet)
        count = self.count

        if self.is_xlwings:
            # calculate new index
            index = sheet.index + offset
            if index < 0:
                index = 0

            # move
            if index >= count - 1:
                after = self.sheets[count-1].sheet_obj.api
                sheet.sheet_obj.api.Move(None, after)
            elif 0 <= index < count:
                before = self.sheets[index].sheet_obj.api
                sheet.sheet_obj.api.Move(before)
        else:
            self.book_obj.move_sheet(sheet.sheet_obj, offset=offset)

        return self

    def copy_sheet(self, from_sheet) -> WorkSheet:
        """
        Copy sheet , and create a new worksheet.

        :Chinese: 复制 WorkSheet, 创建一个新的WorkSheet

        :param from_sheet: the Worksheet to be copied
        :return: return the new-created Worksheet object
        """
        if isinstance(from_sheet, str):
            from_sheet = self._get_sheet(from_sheet)

        if isinstance(from_sheet, WorkSheet):
            from_sheet = from_sheet.sheet_obj

        ws = self.book_obj.copy_worksheet(from_sheet)
        return WorkSheet(self, ws)

    def delete_sheet(self, sheet):
        """
        Delete a worksheet.

        :param sheet: the worksheet name or object
        :return: self
        """
        sheet = self._get_sheet(sheet, False)
        if self.is_xlwings:
            sheet.sheet_obj.delete()
        else:
            self.book_obj.remove(sheet.sheet_obj)

    def index(self, sheet):
        """ retunn the index (int) of the worksheet """
        # noinspection PyBroadException
        try:
            sheet = self._get_sheet(sheet)
            return self.book_obj.index(sheet.sheet_obj)
        except Exception:
            return -1

    @property
    def active(self) -> WorkSheet:
        """ return current activated worksheet """
        if self.is_xlwings:
            return WorkSheet(self, self.book_obj.sheets.active)
        else:
            return WorkSheet(self, self.book_obj.active)

    @active.setter
    def active(self, sheet):
        """ set current activated worksheet """
        sheet = self._get_sheet(sheet)
        if self.is_xlwings:
            sheet.sheet_obj.activate()
        else:
            self.book_obj.active = sheet.sheet_obj

    def protect(self, password: str):
        """
        protect the file.

        https://openpyxl.readthedocs.io/en/stable/protection.html

        :param password:  the password
        :return: self
        """
        if self.is_xlwings:
            # https://docs.microsoft.com/en-us/office/vba/api/excel.workbook.protect
            self.book_obj.api.Protect(password, True)
            self.active['A12'].value = password
        else:
            self._obj.security.workbookPassword = password
            self._obj.security.lockStructure = True
        return self

    def unprotect(self, password: str):
        """
        unprotect the file.

        :param password: the password
        :return: self
        """
        if self.is_xlwings:
            # https://docs.microsoft.com/en-us/office/vba/api/excel.workbook.unprotect
            self.book_obj.api.Unprotect(password)
        else:
            self._obj.security.workbookPassword = password
            self._obj.security.lockStructure = False
        return self

    def set_file_password(self, password: str, hashed=False):
        """ set file password """
        if self.is_xlwings:
            self._file_password = password
        else:
            self._obj.security.set_workbook_password(password, already_hashed=hashed)
            return True
        return False

    # noinspection PyMethodMayBeStatic
    def _parse_item(self, item):
        """
        parese the item of __getitem___

        :param item:
        :return: sheet_name, address, row, col, end_row, end_col
        """
        sheet_name = None
        address = None
        row = None
        col = None
        end_row = None
        end_col = None

        if isinstance(item, tuple):
            for n in item:
                if isinstance(n, str):
                    if sheet_name is None:
                        sheet_name = n
                    elif address is None:
                        address = n
                if isinstance(n, int):
                    if row is None:
                        row = n
                    elif col is None:
                        col = n
                    elif end_row is None:
                        end_row = n
                    elif end_col is None:
                        end_col = n
        elif isinstance(item, str):
            offset = item.find('!')
            if offset >= 0:
                sheet_name = item[:offset]
                address = item[offset+1:]
            else:
                sheet_name = item

        return sheet_name, address, row, col, end_row, end_col

    def __getitem__(self, item):
        if isinstance(item, int):
            return self._get_sheet_by_index(item)

        sheet_name, address, row, col, end_row, end_col = self._parse_item(item)

        if sheet_name is None:
            raise ValueError("sheet name should specified")

        if sheet_name not in self:
            self.create_sheet(sheet_name)

        if sheet_name in self:
            sheet = self._get_sheet(sheet_name)
            if address is not None:
                return sheet[address]
            elif row is None and col is None:
                return sheet
            elif end_row is None and end_col is None:
                return sheet.cell(row, col)
            else:
                end_row = row if end_row is None else end_row
                end_col = col if end_col is None else end_col
                addr = [row, col, end_row, end_col]
                return Range(addr, sheet)

    def __repr__(self):
        return '<Excel WorkBook %s>' % repr(self.filename)

    def print(self, **kwargs):
        """
        Call Excel Application,  print the file to printer.

        :Chinese: 调用Excel, 使用打印机打印文档

        :param kwargs:
        :return: self
        """
        print("printing", self.filename, '...')

        if self.is_xlwings:
            self.book_obj.api.PrintOut(**kwargs)
        else:
            app, workbook = Win32.open_excel(self._absolute_filename, False)
            workbook.PrintOut(**kwargs)
            time.sleep(2)
            Win32.close(app, workbook)
        return self

    def find(self, sheet_name, value1, value2=None, regex=False):
        """
        find a cell in specified sheet

        :param sheet_name: worksheet name
        :param value1:   first value
        :param value2:  (optional) second value
        :param regex:   (optional) whether use regex

        :return: return Cell object if found. return None if not found.
        """
        if sheet_name in self:
            sheet = self._get_sheet(sheet_name)
            return sheet.find(value1, value2, regex)
